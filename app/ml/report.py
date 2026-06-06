"""
PDF & Excel report generator for all four report types.
"""
import os
import io
from datetime import datetime

# ── PDF ───────────────────────────────────────────────────────────────────────
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                Table, TableStyle, HRFlowable)
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.units import inch
from flask import current_app

# ── Excel ─────────────────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

PRIMARY   = colors.HexColor('#1E3A5F')
SECONDARY = colors.HexColor('#2E86C1')
ACCENT    = colors.HexColor('#E67E22')
LIGHT     = colors.HexColor('#EBF5FB')
WHITE     = colors.white

# ── Shared Helpers ────────────────────────────────────────────────────────────

def shared_draw_header(canvas, doc):
    canvas.saveState()
    logo_path = os.path.join(current_app.root_path, 'static', 'img', 'logo.png')
    page_w, page_h = doc.pagesize
    
    if os.path.exists(logo_path):
        canvas.drawImage(logo_path, 40, page_h - 60, width=1.8*inch, height=0.6*inch, preserveAspectRatio=True, anchor='nw')
        
    canvas.setFont('Helvetica-Bold', 16)
    canvas.drawString(180, page_h - 35, "PT Putra Hamid")
    
    canvas.setFont('Helvetica', 10)
    canvas.setFillColor(colors.dimgrey)
    canvas.drawString(180, page_h - 50, "Sistem Prediksi Harga Motor Bekas")
    
    canvas.setStrokeColor(colors.grey)
    canvas.setLineWidth(1)
    canvas.line(40, page_h - 70, page_w - 40, page_h - 70)
    
    canvas.setFont('Helvetica', 9)
    canvas.setFillColor(colors.grey)
    canvas.drawRightString(page_w - 40, 20, f"Halaman {doc.page}")
    
    canvas.restoreState()


def _header(story, styles, title, subtitle=''):
    story.append(Paragraph(title, styles['h1']))
    if subtitle:
        story.append(Paragraph(subtitle, styles['sub']))
    story.append(Spacer(1, 0.3*cm))


def _make_styles():
    s = getSampleStyleSheet()

    def _add(name, **kwargs):
        if name not in s:
            s.add(ParagraphStyle(name, **kwargs))

    _add('title', fontSize=14, alignment=TA_CENTER, fontName='Helvetica-Bold')
    _add('sub',   fontSize=9,  alignment=TA_CENTER, textColor=colors.grey)
    _add('h1',    fontSize=12, fontName='Helvetica-Bold', textColor=PRIMARY)
    _add('body',  fontSize=9,  leading=14)
    return s


def _table_style():
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), PRIMARY),
        ('TEXTCOLOR',  (0, 0), (-1, 0), WHITE),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, 0), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [WHITE, LIGHT]),
        ('GRID',       (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ('FONTSIZE',   (0, 1), (-1, -1), 8),
        ('LEFTPADDING',  (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING',   (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 4),
    ])


def _rp(value):
    return f"Rp {int(value):,}".replace(',', '.')


# ── 1. Dataset Statistics Report ─────────────────────────────────────────────

def generate_dataset_pdf(dataset, stats, save_path):
    doc  = SimpleDocTemplate(save_path, pagesize=A4,
                              topMargin=90, bottomMargin=40,
                              leftMargin=2*cm, rightMargin=2*cm)
    s    = _make_styles()
    story = []

    _header(story, s, 'Laporan Statistik Dataset',
            f'File: {dataset.nama_file}  |  Dicetak: {datetime.now():%d %B %Y %H:%M}')

    # Summary
    story.append(Paragraph('<b>Ringkasan Dataset</b>', s['h1']))
    story.append(Spacer(1, 0.2*cm))
    summary_data = [
        ['Informasi', 'Nilai'],
        ['Nama File',      dataset.nama_file],
        ['Jumlah Data',    str(dataset.jumlah_data)],
        ['Status',         dataset.status.capitalize()],
        ['Tanggal Upload', dataset.created_at.strftime('%d %B %Y %H:%M')],
    ]
    t = Table(summary_data, colWidths=[6*cm, 10*cm])
    t.setStyle(_table_style())
    story.append(t)
    story.append(Spacer(1, 0.5*cm))

    # Numerical stats
    num_map = {'tahun_produksi': 'Tahun Produksi', 'kapasitas_mesin': 'Kapasitas Mesin (cc)',
               'jarak_tempuh': 'Jarak Tempuh (km)', 'harga_aktual': 'Harga Aktual (Rp)'}
    story.append(Paragraph('<b>Statistik Deskriptif Numerik</b>', s['h1']))
    story.append(Spacer(1, 0.2*cm))
    num_data = [['Variabel', 'Min', 'Max', 'Mean', 'Median', 'Std Dev']]
    for col, label in num_map.items():
        if col in stats:
            st = stats[col]
            num_data.append([
                label,
                f"{st['min']:,.0f}", f"{st['max']:,.0f}",
                f"{st['mean']:,.2f}", f"{st['median']:,.0f}",
                f"{st['std']:,.2f}",
            ])
    t2 = Table(num_data, colWidths=[4.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm])
    t2.setStyle(_table_style())
    story.append(t2)
    story.append(Spacer(1, 0.5*cm))

    doc.build(story, onFirstPage=shared_draw_header, onLaterPages=shared_draw_header)
    return save_path


def generate_dataset_excel(dataset, stats, save_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Statistik Dataset'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1E3A5F')

    ws['A1'] = 'Laporan Statistik Dataset — PT Putra Hamid'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')

    headers = ['Variabel', 'Min', 'Max', 'Mean', 'Median', 'Std Dev']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')

    num_map = {'tahun_produksi': 'Tahun Produksi', 'kapasitas_mesin': 'Kapasitas Mesin (cc)',
               'jarak_tempuh': 'Jarak Tempuh (km)', 'harga_aktual': 'Harga Aktual (Rp)'}
    row = 4
    for col, label in num_map.items():
        if col in stats:
            st = stats[col]
            ws.append([label, st['min'], st['max'], st['mean'], st['median'], st['std']])
            row += 1

    for i in range(1, 7):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = 18

    wb.save(save_path)
    return save_path


# ── 2. Coefficient Report ────────────────────────────────────────────────────

def generate_koefisien_pdf(model_obj, save_path):
    doc   = SimpleDocTemplate(save_path, pagesize=A4,
                               topMargin=90, bottomMargin=40,
                               leftMargin=2*cm, rightMargin=2*cm)
    s     = _make_styles()
    story = []

    _header(story, s, 'Laporan Koefisien Regresi',
            f'Model ID: {model_obj.id}  |  Dicetak: {datetime.now():%d %B %Y %H:%M}')

    # Equation
    story.append(Paragraph('<b>Persamaan Regresi Linear Berganda</b>', s['h1']))
    story.append(Spacer(1, 0.2*cm))
    eq = f"Ŷ = {model_obj.intercept:.2f}"
    for name, coef in (model_obj.koefisien or {}).items():
        sign = '+' if coef >= 0 else '-'
        eq += f" {sign} {abs(coef):.4f}·{name}"
    story.append(Paragraph(eq, s['body']))
    story.append(Spacer(1, 0.4*cm))

    # Coefficients table
    story.append(Paragraph('<b>Tabel Koefisien</b>', s['h1']))
    story.append(Spacer(1, 0.2*cm))
    coef_data = [['Variabel', 'Koefisien (β)', 'Pengaruh']]
    for name, coef in (model_obj.koefisien or {}).items():
        arah = '↑ Positif' if coef > 0 else '↓ Negatif'
        coef_data.append([name, f"{coef:.4f}", arah])
    t = Table(coef_data, colWidths=[8*cm, 5*cm, 4*cm])
    t.setStyle(_table_style())
    story.append(t)
    story.append(Spacer(1, 0.5*cm))

    # Metrics
    story.append(Paragraph('<b>Metrik Model</b>', s['h1']))
    story.append(Spacer(1, 0.2*cm))
    metric_data = [
        ['Metrik', 'Nilai'],
        ['R² Score', f"{model_obj.r2_score:.4f}"],
        ['MAE',      _rp(model_obj.mae)],
        ['RMSE',     _rp(model_obj.rmse)],
        ['MAPE',     f"{model_obj.mape:.2f}%"],
        ['Training Size', str(model_obj.train_size)],
        ['Test Size',     str(model_obj.test_size)],
    ]
    t2 = Table(metric_data, colWidths=[7*cm, 10*cm])
    t2.setStyle(_table_style())
    story.append(t2)

    doc.build(story, onFirstPage=shared_draw_header, onLaterPages=shared_draw_header)
    return save_path


# ── 3. Prediction Result Report ──────────────────────────────────────────────

def generate_prediksi_pdf(prediksi_obj, admin_name, save_path):
    doc   = SimpleDocTemplate(save_path, pagesize=A4,
                               topMargin=90, bottomMargin=40,
                               leftMargin=2*cm, rightMargin=2*cm)
    s     = _make_styles()
    story = []

    _header(story, s, 'Laporan Hasil Prediksi Harga Motor',
            f'Admin: {admin_name}  |  Dicetak: {datetime.now():%d %B %Y %H:%M}')

    p = prediksi_obj
    data = [
        ['Parameter', 'Nilai'],
        ['Tahun Produksi',    str(p.tahun_produksi)],
        ['Kapasitas Mesin',   f"{p.kapasitas_mesin} cc"],
        ['Jarak Tempuh',      f"{p.jarak_tempuh:,.0f} km"],
        ['Merek Motor',       p.merek_motor],
        ['Kondisi Fisik',     p.kondisi_fisik],
        ['Tipe Motor',        p.tipe_motor],
        ['Kelengkapan Surat', p.kelengkapan_surat],
        ['Pajak',             p.pajak],
        ['', ''],
        ['HARGA PREDIKSI',   _rp(p.harga_prediksi)],
    ]
    if p.harga_aktual:
        data.append(['Harga Aktual',  _rp(p.harga_aktual)])
        selisih = abs(p.harga_prediksi - p.harga_aktual)
        data.append(['Selisih (Error)', _rp(selisih)])

    t = Table(data, colWidths=[7*cm, 10*cm])
    style = _table_style()
    style.add('FONTNAME',   (0, 10), (-1, 10), 'Helvetica-Bold')
    style.add('BACKGROUND', (0, 10), (-1, 10), ACCENT)
    style.add('TEXTCOLOR',  (0, 10), (-1, 10), WHITE)
    t.setStyle(style)
    story.append(t)

    doc.build(story, onFirstPage=shared_draw_header, onLaterPages=shared_draw_header)
    return save_path


# ── 4. Model Evaluation Report ───────────────────────────────────────────────

def generate_evaluasi_pdf(model_obj, save_path):
    doc   = SimpleDocTemplate(save_path, pagesize=A4,
                               topMargin=90, bottomMargin=40,
                               leftMargin=2*cm, rightMargin=2*cm)
    s     = _make_styles()
    story = []

    _header(story, s, 'Laporan Evaluasi Model',
            f'Model ID: {model_obj.id}  |  Dicetak: {datetime.now():%d %B %Y %H:%M}')

    def interpret_r2(r2):
        if r2 >= 0.9:  return 'Sangat Baik (≥ 0.90)'
        if r2 >= 0.7:  return 'Baik (0.70 – 0.89)'
        if r2 >= 0.5:  return 'Cukup (0.50 – 0.69)'
        return 'Kurang Baik (< 0.50)'

    metric_data = [
        ['Metrik', 'Nilai', 'Interpretasi'],
        ['R² (R-Squared)', f"{model_obj.r2_score:.4f}", interpret_r2(model_obj.r2_score)],
        ['MAE',  _rp(model_obj.mae),  'Rata-rata error absolut'],
        ['MSE',  _rp(model_obj.mse),  'Rata-rata kuadrat error'],
        ['RMSE', _rp(model_obj.rmse), 'Akar rata-rata kuadrat error'],
        ['MAPE', f"{model_obj.mape:.2f}%", 'Rata-rata persentase error'],
    ]
    t = Table(metric_data, colWidths=[4*cm, 6*cm, 7*cm])
    t.setStyle(_table_style())
    story.append(t)
    story.append(Spacer(1, 0.5*cm))

    # Comparison table (first 20 test rows)
    y_test = model_obj.y_test_data or []
    y_pred = model_obj.y_pred_data or []
    story.append(Paragraph('<b>Perbandingan Harga Aktual vs Prediksi (Data Test)</b>', s['h1']))
    story.append(Spacer(1, 0.2*cm))
    cmp_data = [['No', 'Harga Aktual (Rp)', 'Harga Prediksi (Rp)', 'Selisih (Rp)']]
    for i, (ya, yp) in enumerate(zip(y_test[:20], y_pred[:20]), 1):
        cmp_data.append([str(i), _rp(ya), _rp(yp), _rp(abs(ya - yp))])
    t2 = Table(cmp_data, colWidths=[1.5*cm, 5*cm, 5*cm, 5*cm])
    t2.setStyle(_table_style())
    story.append(t2)

    doc.build(story, onFirstPage=shared_draw_header, onLaterPages=shared_draw_header)
    return save_path


def generate_evaluasi_excel(model_obj, save_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Evaluasi Model'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1E3A5F')

    ws['A1'] = 'Laporan Evaluasi Model — PT Putra Hamid'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')

    metrics = [
        ['Metrik', 'Nilai'],
        ['R²',   model_obj.r2_score],
        ['MAE',  model_obj.mae],
        ['MSE',  model_obj.mse],
        ['RMSE', model_obj.rmse],
        ['MAPE', model_obj.mape],
    ]
    for r, row in enumerate(metrics, 3):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if r == 3:
                cell.font = header_font
                cell.fill = header_fill

    # Test data
    y_test = model_obj.y_test_data or []
    y_pred = model_obj.y_pred_data or []
    start_row = 11
    for c, h in enumerate(['No', 'Aktual', 'Prediksi', 'Selisih'], 1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for i, (ya, yp) in enumerate(zip(y_test, y_pred), 1):
        ws.append([i, ya, yp, abs(ya - yp)])

    for i in range(1, 5):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = 20
    wb.save(save_path)
    return save_path


def generate_semua_prediksi_pdf(results, admin_name, save_path, merek_f='', tipe_f=''):
    """
    Generate landscape A4 PDF report of all prediction records, optionally filtered.
    """
    doc   = SimpleDocTemplate(save_path, pagesize=landscape(A4),
                               topMargin=90, bottomMargin=40,
                               leftMargin=1.5*cm, rightMargin=1.5*cm)
    s     = _make_styles()
    story = []

    # Filter details
    filter_details = []
    if merek_f:
        filter_details.append(f"Merek: {merek_f}")
    if tipe_f:
        filter_details.append(f"Tipe: {tipe_f}")

    subtitle = f"Admin: {admin_name}  |  Dicetak: {datetime.now():%d %B %Y %H:%M}"
    if filter_details:
        subtitle += f"  |  Filter: {', '.join(filter_details)}"
    else:
        subtitle += "  |  Filter: Semua Data"
    
    subtitle += f"  |  Total: {len(results)} Data"

    _header(story, s, 'Laporan Riwayat Hasil Prediksi Harga Motor Bekas', subtitle)
    story.append(Spacer(1, 0.4*cm))

    # Columns: No, Merek, Tipe, Model, Tahun, KM, Kondisi, Pajak, Surat, Harga Prediksi, Tanggal
    headers = [
        'No', 'Merek', 'Tipe', 'Model', 'Tahun', 'Jarak Tempuh',
        'Kondisi', 'Pajak', 'Surat', 'Harga Prediksi', 'Tanggal'
    ]
    table_data = [headers]

    for i, r in enumerate(results, 1):
        formatted_km = f"{r.jarak_tempuh:,.0f}".replace(',', '.') + " km"
        formatted_harga = _rp(r.harga_prediksi)
        formatted_date = r.created_at.strftime('%d/%m/%Y')
        
        table_data.append([
            str(i),
            r.merek_motor,
            r.tipe_motor,
            r.model_motor or '-',
            str(r.tahun_produksi),
            formatted_km,
            r.kondisi_fisik,
            r.pajak,
            r.kelengkapan_surat,
            formatted_harga,
            formatted_date
        ])

    # A4 Landscape width is 29.7cm. Margins are 1.5cm left/right. Printable width is 26.7cm.
    # Total widths = 1.0 + 2.5 + 2.5 + 3.5 + 1.5 + 2.5 + 2.0 + 1.8 + 3.2 + 3.5 + 2.5 = 26.5 cm
    col_widths = [
        1.0*cm,  # No
        2.5*cm,  # Merek
        2.5*cm,  # Tipe
        3.5*cm,  # Model
        1.5*cm,  # Tahun
        2.5*cm,  # KM
        2.0*cm,  # Kondisi
        1.8*cm,  # Pajak
        3.2*cm,  # Surat
        3.5*cm,  # Harga Prediksi
        2.5*cm   # Tanggal
    ]

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(_table_style())
    story.append(t)

    doc.build(story, onFirstPage=shared_draw_header, onLaterPages=shared_draw_header)
    return save_path

